import io
import re
import calendar
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Column spec: (HEADER, field, width). "_total_amount" is computed.
# The WAREHOUSE column is inserted only for sheets that contain Amazon rows
# (Flipkart / Other platforms have no warehouse) — see _columns().
_BASE_COLS = [
    ("QTY",           "qty",               6),
    ("TYPE",          "transaction_type",  10),
    ("PARTY NAME",    "party_name",        28),
    ("GST NUMBER",    "gst_number",        18),
    ("INV NO",        "inv_no",            28),
    ("INV DATE",      "inv_date",          12),
    ("TAXABLE VALUE", "taxable_value",     16),
    ("CGST9",         "cgst9",             12),
    ("SGST9",         "sgst9",             12),
    ("IGST18",        "igst18",            12),
    ("TOTAL AMOUNT",  "_total_amount",     16),
    ("PARTY ADDRESS", "party_address",     22),
]
_WAREHOUSE_COL = ("WAREHOUSE", "warehouse", 14)
_NUM_HEADERS = {"TAXABLE VALUE", "CGST9", "SGST9", "IGST18", "TOTAL AMOUNT"}


def _columns(show_warehouse: bool):
    """Return (headers, fields, widths, num_cols_1indexed) for a sheet.
    WAREHOUSE is slotted right after TYPE when requested."""
    cols = list(_BASE_COLS)
    if show_warehouse:
        cols.insert(2, _WAREHOUSE_COL)   # after QTY, TYPE
    headers = [c[0] for c in cols]
    fields  = [c[1] for c in cols]
    widths  = [c[2] for c in cols]
    num_cols = {i for i, h in enumerate(headers, start=1) if h in _NUM_HEADERS}
    return headers, fields, widths, num_cols


TITLE_FONT   = Font(bold=True, size=13)
HEADER_FONT  = Font(bold=True, size=11)
HEADER_FILL  = PatternFill("solid", fgColor="D9E1F2")
SALE_FILL    = PatternFill("solid", fgColor="EBF5EB")
RETURN_FILL  = PatternFill("solid", fgColor="FFF0F0")
CANCEL_FONT  = Font(color="FF0000", italic=True)
TOTAL_FONT   = Font(bold=True, size=11)
TOTAL_FILL   = PatternFill("solid", fgColor="FFF2CC")
THIN         = Side(style="thin", color="BFBFBF")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

AMAZON_WAREHOUSES   = ["IN", "MAA4", "CJB1"]
WAREHOUSE_LABELS    = {"IN": "IN (India)", "MAA4": "MAA4 (Chennai)", "CJB1": "CJB1 (Coimbatore)"}


def _sort_key(inv):
    v = (inv.inv_no or "").strip()
    return v.upper()


def _month_key(inv_date) -> str:
    """Parse any OCR date format → 'YYYY-MM', else 'Unknown'.
    Same logic as routes/stats.py — handles DD.MM.YYYY, DD/MM/YY,
    DD-MM-YYYY, 'DD-MM-YYYY, HH:MM AM', DD-Mon-YY, etc."""
    if not inv_date:
        return "Unknown"
    s = str(inv_date).strip()
    s = re.split(r",\s*\d", s)[0].strip()  # strip time portion

    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%d %b %y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m")
        except ValueError:
            pass

    for sep in (".", "/", "-"):
        if sep not in s:
            continue
        parts = [p.strip() for p in s.split(sep)]
        if len(parts) != 3:
            continue
        d, m, y = parts[0], parts[1], parts[2]
        if len(y) == 2 and y.isdigit():
            y = "20" + y
        if len(y) == 4 and y.isdigit() and m.isdigit() and 1 <= int(m) <= 12:
            return f"{y}-{m.zfill(2)}"
        break

    return "Unknown"


def _month_label(key: str) -> str:
    """'2026-04' → 'Apr 2026'; 'Unknown' → 'Undated'."""
    if key == "Unknown":
        return "Undated"
    y, m = key.split("-")
    return f"{calendar.month_abbr[int(m)]} {y}"


def _write_sheet(ws, invoices, title: str, show_warehouse: bool = False):
    headers, fields, widths, num_cols = _columns(show_warehouse)

    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 6

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 6

    type_col = headers.index("TYPE") + 1

    data_start = 5
    sorted_invs = sorted(invoices, key=_sort_key)
    for row_idx, inv in enumerate(sorted_invs, start=data_start):
        is_return = getattr(inv, "transaction_type", "Sale") == "Return"
        is_cancel = getattr(inv, "cancelled", False)
        row_fill  = RETURN_FILL if is_return else SALE_FILL

        for col_idx, field in enumerate(fields, start=1):
            if field == "_total_amount":
                tv = getattr(inv, "taxable_value", None) or 0
                cg = getattr(inv, "cgst9", None) or 0
                sg = getattr(inv, "sgst9", None) or 0
                ig = getattr(inv, "igst18", None) or 0
                val = tv + cg + sg + ig if (tv or cg or sg or ig) else None
            else:
                val = getattr(inv, field, None)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if not is_cancel:
                cell.fill = row_fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center",
                                       horizontal="right" if col_idx in num_cols else "left")
            if is_cancel:
                cell.font = CANCEL_FONT
            elif is_return and col_idx == type_col:
                cell.font = Font(bold=True, color="CC0000")
            if col_idx in num_cols:
                cell.number_format = "#,##0.00"

        if is_cancel:
            ws.cell(row=row_idx, column=len(headers) + 1, value="CANCEL").font = CANCEL_FONT

    data_end = data_start + len(sorted_invs) - 1

    if sorted_invs:
        tr = data_end + 2
        ws.cell(row=tr, column=1, value="TOTAL").font = TOTAL_FONT
        ws.cell(row=tr, column=1).fill = TOTAL_FILL
        ws.cell(row=tr, column=1).alignment = Alignment(horizontal="center")
        for col_idx in num_cols:
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=tr, column=col_idx,
                           value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})")
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = BORDER

    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = f"A{data_start}"


def build_excel(invoices) -> bytes:
    wb = Workbook()
    first = True

    def _add_sheet(title, rows, header):
        nonlocal first
        if not rows:
            return
        ws = wb.active if first else wb.create_sheet(title)
        if first:
            ws.title = title
            first = False
        # WAREHOUSE column only when the sheet actually holds Amazon invoices
        show_wh = any((getattr(r, "platform", "") or "") == "Amazon" for r in rows)
        _write_sheet(ws, rows, header, show_warehouse=show_wh)

    # ── Group data ───────────────────────────────────────────────────────
    by_platform: dict = defaultdict(list)
    by_platform_type: dict = defaultdict(lambda: defaultdict(list))
    # Amazon: further split by warehouse
    amazon_by_wh: dict = defaultdict(list)
    amazon_by_wh_type: dict = defaultdict(lambda: defaultdict(list))

    for inv in invoices:
        platform = (inv.platform or "Other").strip()
        t = getattr(inv, "transaction_type", "Sale") or "Sale"
        is_cancel = getattr(inv, "cancelled", False)
        by_platform[platform].append(inv)
        # Dedicated Sales / Returns sheets must hold ONLY live invoices —
        # cancelled ones are excluded here (they still appear, marked CANCEL,
        # in the combined "all invoices" sheets below).
        if not is_cancel:
            by_platform_type[platform][t].append(inv)
        if platform == "Amazon":
            wh = (inv.warehouse or "IN").upper()
            amazon_by_wh[wh].append(inv)
            if not is_cancel:
                amazon_by_wh_type[wh][t].append(inv)

    # ── Flipkart sheets ──────────────────────────────────────────────────
    if "Flipkart" in by_platform:
        _add_sheet("Flipkart Sales",   by_platform_type["Flipkart"]["Sale"],   "FLIPKART — SALES")
        _add_sheet("Flipkart Returns", by_platform_type["Flipkart"]["Return"], "FLIPKART — RETURNS")
        _add_sheet("Flipkart",         by_platform["Flipkart"],                "FLIPKART — ALL INVOICES")

    # ── Amazon per-warehouse sheets ──────────────────────────────────────
    if "Amazon" in by_platform:
        for wh in AMAZON_WAREHOUSES:
            if wh not in amazon_by_wh:
                continue
            label = WAREHOUSE_LABELS.get(wh, wh)
            _add_sheet(f"Amazon {wh} Sales",   amazon_by_wh_type[wh]["Sale"],   f"AMAZON {label} — SALES")
            _add_sheet(f"Amazon {wh} Returns", amazon_by_wh_type[wh]["Return"], f"AMAZON {label} — RETURNS")
            _add_sheet(f"Amazon {wh}",         amazon_by_wh[wh],                f"AMAZON {label} — ALL INVOICES")
        # Combined Amazon sheet
        _add_sheet("Amazon",  by_platform["Amazon"],  "AMAZON — ALL INVOICES")

    # ── Other platforms ──────────────────────────────────────────────────
    for p in by_platform:
        if p in ("Flipkart", "Amazon"):
            continue
        _add_sheet(f"{p} Sales",   by_platform_type[p]["Sale"],   f"{p.upper()} — SALES")
        _add_sheet(f"{p} Returns", by_platform_type[p]["Return"], f"{p.upper()} — RETURNS")
        _add_sheet(p,              by_platform[p],                f"{p.upper()} — ALL INVOICES")

    # ── ALL combined ─────────────────────────────────────────────────────
    _add_sheet("ALL", list(invoices), "ALL INVOICES")

    # ── Month-wise sheets — full detail, every row, split Sales vs Returns ─
    by_month: dict = defaultdict(list)
    for inv in invoices:
        by_month[_month_key(inv.inv_date)].append(inv)

    month_keys = sorted(k for k in by_month if k != "Unknown")
    if "Unknown" in by_month:
        month_keys.append("Unknown")

    # One Sales sheet + one Returns sheet per month, holding every individual
    # invoice row (cancelled excluded). Skipped when only a single month is
    # present — the platform Sales/Returns sheets already cover that case.
    if len(month_keys) > 1:
        for mk in month_keys:
            label = _month_label(mk)
            rows = [r for r in by_month[mk] if not getattr(r, "cancelled", False)]
            sale_rows   = [r for r in rows if (r.transaction_type or "Sale") != "Return"]
            return_rows = [r for r in rows if (r.transaction_type or "Sale") == "Return"]
            _add_sheet(f"{label} Sales",   sale_rows,   f"{label.upper()} — SALES")
            _add_sheet(f"{label} Returns", return_rows, f"{label.upper()} — RETURNS")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
